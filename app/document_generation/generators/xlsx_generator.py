import io
import re
from typing import Dict, Any, Optional, List
from html.parser import HTMLParser
from app.document_generation.generators.base import BaseGenerator


class _XLSXTableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.tables = []
        self.current_table = {"headers": [], "rows": [], "title": ""}
        self.current_row = []
        self.in_table = False
        self.in_header = False
        self.in_row = False
        self.cell_text = ""
        self.current_title = ""
        self.in_title = False
        self.titles = []

    def handle_starttag(self, tag, attrs):
        tag = tag.lower()
        if tag == "table":
            if self.current_table["rows"] or self.current_table["headers"]:
                self._finalize_table()
            self.in_table = True
            self.current_table = {"headers": [], "rows": [], "title": self.current_title}
        elif tag == "tr":
            self.current_row = []
            self.in_row = True
        elif tag == "th":
            self.in_header = True
            self.cell_text = ""
        elif tag == "td":
            self.in_header = False
            self.cell_text = ""
        elif tag in ("h1", "h2", "h3", "h4"):
            self.in_title = True
            self.cell_text = ""

    def handle_endtag(self, tag):
        tag = tag.lower()
        if tag == "table":
            self._finalize_table()
            self.in_table = False
        elif tag == "tr":
            if self.current_row:
                if self.current_table["headers"] and not self.in_header:
                    self.current_table["rows"].append(list(self.current_row))
                elif self.in_header or (self.current_table["headers"] and False):
                    pass
                else:
                    self.current_table["rows"].append(list(self.current_row))
            self.current_row = []
            self.in_row = False
        elif tag in ("th",):
            text = self.cell_text.strip()
            self.current_table["headers"].append(text)
            self.in_header = False
            self.cell_text = ""
        elif tag == "td":
            text = self.cell_text.strip()
            self.current_row.append(text)
            self.cell_text = ""
        elif tag in ("h1", "h2", "h3", "h4"):
            self.current_title = self.cell_text.strip()
            self.in_title = False
            self.cell_text = ""

    def handle_data(self, data):
        if self.in_title:
            self.cell_text += data
        elif self.in_table and (self.in_header or self.in_row):
            self.cell_text += data

    def _finalize_table(self):
        headers = self.current_table.get("headers", [])
        rows = self.current_table.get("rows", [])
        if headers or rows:
            self.tables.append({
                "title": self.current_table.get("title", ""),
                "headers": headers,
                "rows": rows,
            })
        self.current_table = {"headers": [], "rows": [], "title": ""}


class _XLSXContentParser(HTMLParser):
    """Parse non-table HTML into structured rows for Excel output."""
    def __init__(self):
        super().__init__()
        self.rows = []       # list of {"type": ..., "text": ...}
        self.text_buf = ""
        self.tag_stack = []
        self.in_li = False
        self.in_code = False
        self.code_buf = ""

    def handle_starttag(self, tag, attrs):
        tag = tag.lower()
        self.tag_stack.append(tag)
        if tag in ("h1", "h2", "h3", "h4", "h5", "h6"):
            self._flush()
        elif tag == "li":
            self._flush()
            self.in_li = True
        elif tag in ("pre", "code"):
            self._flush()
            self.in_code = True
            self.code_buf = ""
        elif tag == "br":
            self.text_buf += "\n"
        elif tag == "hr":
            self._flush()
            self.rows.append({"type": "separator", "text": ""})
        elif tag == "p":
            self._flush()

    def handle_endtag(self, tag):
        tag = tag.lower()
        if self.tag_stack and self.tag_stack[-1] == tag:
            self.tag_stack.pop()
        if tag in ("h1", "h2", "h3", "h4", "h5", "h6"):
            text = self.text_buf.strip()
            if text:
                self.rows.append({"type": "heading", "text": text, "level": int(tag[1])})
            self.text_buf = ""
        elif tag == "li":
            text = self.text_buf.strip()
            if text:
                self.rows.append({"type": "list_item", "text": text})
            self.text_buf = ""
            self.in_li = False
        elif tag in ("pre", "code"):
            if self.code_buf.strip():
                self.rows.append({"type": "code", "text": self.code_buf.strip()})
            self.code_buf = ""
            self.in_code = False
        elif tag == "p":
            self._flush()

    def handle_data(self, data):
        if self.in_code:
            self.code_buf += data
        else:
            self.text_buf += data

    def _flush(self):
        text = self.text_buf.strip()
        if text:
            self.rows.append({"type": "text", "text": text})
        self.text_buf = ""

    def finalize(self):
        self._flush()
        return self.rows


class XLSXGenerator(BaseGenerator):
    @property
    def format(self) -> str:
        return "xlsx"

    @property
    def mime_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    @property
    def file_extension(self) -> str:
        return ".xlsx"

    def generate(self, html: str, template: Optional[Dict[str, Any]] = None, metadata: Optional[Dict[str, Any]] = None) -> bytes:
        try:
            return self._generate_with_openpyxl(html)
        except ImportError:
            raise ImportError("openpyxl is required for XLSX generation.")

    def generate_preview(self, html: str, template: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        parser = _XLSXTableParser()
        parser.feed(html)

        preview = {
            "sheets": [],
        }

        if parser.tables:
            for table in parser.tables:
                sheet_data = {
                    "name": table["title"] or "Sheet",
                    "headers": table["headers"],
                    "rows": table["rows"][:100],
                    "total_rows": len(table["rows"]),
                }
                preview["sheets"].append(sheet_data)
        else:
            text = re.sub(r"<[^>]+>", "", html).strip()[:500]
            preview["sheets"].append({
                "name": "Content",
                "headers": ["Content"],
                "rows": [[line] for line in text.split("\n") if line.strip()],
                "total_rows": len(text.split("\n")),
            })

        return preview

    def _generate_with_openpyxl(self, html: str) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        wb.remove(wb.active)

        # First try to extract HTML tables
        table_parser = _XLSXTableParser()
        table_parser.feed(html)

        if table_parser.tables:
            for idx, table in enumerate(table_parser.tables):
                title = table.get("title", f"Sheet {idx + 1}")[:31]
                ws = wb.create_sheet(title=title)
                self._write_table(ws, table)
        else:
            # Intelligent fallback: parse structured content
            ws = wb.create_sheet(title="Content")
            self._write_structured_content(ws, html)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _write_structured_content(self, ws, html: str):
        """Parse HTML into structured, formatted Excel rows."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        content_parser = _XLSXContentParser()
        content_parser.feed(html)
        items = content_parser.finalize()

        if not items:
            ws.cell(row=1, column=1, value="No content found.")
            return

        # Style definitions
        heading_font = Font(bold=True, color="FFFFFF", size=13)
        heading_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        subheading_font = Font(bold=True, color="1A1A2E", size=11)
        subheading_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
        text_font = Font(size=10, color="333333")
        list_font = Font(size=10, color="333333")
        code_font = Font(name="Courier New", size=9, color="2D2D2D")
        code_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        sep_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )
        wrap_alignment = Alignment(vertical="top", wrap_text=True)

        row_idx = 1

        for item in items:
            item_type = item["type"]
            text = item.get("text", "")

            if item_type == "heading":
                level = item.get("level", 1)
                cell = ws.cell(row=row_idx, column=1, value=text)
                if level <= 2:
                    cell.font = heading_font
                    cell.fill = heading_fill
                else:
                    cell.font = subheading_font
                    cell.fill = subheading_fill
                cell.alignment = Alignment(vertical="center")
                cell.border = thin_border
                # Merge across columns A-B for heading rows
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
                row_idx += 1

            elif item_type == "list_item":
                cell = ws.cell(row=row_idx, column=1, value="•")
                cell.font = list_font
                cell.alignment = Alignment(horizontal="center", vertical="top")
                cell.border = thin_border

                val_cell = ws.cell(row=row_idx, column=2, value=text)
                val_cell.font = list_font
                val_cell.alignment = wrap_alignment
                val_cell.border = thin_border
                row_idx += 1

            elif item_type == "code":
                for code_line in text.split("\n"):
                    cell = ws.cell(row=row_idx, column=1, value=code_line)
                    cell.font = code_font
                    cell.fill = code_fill
                    cell.alignment = wrap_alignment
                    cell.border = thin_border
                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
                    row_idx += 1

            elif item_type == "separator":
                cell = ws.cell(row=row_idx, column=1, value="")
                cell.fill = sep_fill
                cell.border = thin_border
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
                row_idx += 1

            elif item_type == "text":
                # Try to detect key-value pattern: "Key: Value" or "Key — Value"
                kv_match = re.match(r"^(.+?)\s*[:—–]\s+(.+)$", text)
                if kv_match and len(kv_match.group(1)) < 60:
                    key_cell = ws.cell(row=row_idx, column=1, value=kv_match.group(1).strip())
                    key_cell.font = Font(bold=True, size=10, color="333333")
                    key_cell.alignment = Alignment(vertical="top")
                    key_cell.border = thin_border

                    val_cell = ws.cell(row=row_idx, column=2, value=kv_match.group(2).strip())
                    val_cell.font = text_font
                    val_cell.alignment = wrap_alignment
                    val_cell.border = thin_border
                else:
                    cell = ws.cell(row=row_idx, column=1, value=text)
                    cell.font = text_font
                    cell.alignment = wrap_alignment
                    cell.border = thin_border
                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
                row_idx += 1

        # Auto-size columns
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 70

    def _write_table(self, ws, table: Dict):
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        row_idx = 1
        if table.get("headers"):
            for col_idx, header in enumerate(table["headers"], 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            row_idx += 1

        alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        for row_num, row_data in enumerate(table.get("rows", [])):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if row_num % 2 == 1:
                    cell.fill = alt_fill
            row_idx += 1

        num_cols = len(table.get("headers", [])) or 1
        for col_idx in range(1, num_cols + 1):
            max_length = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
                for cell_val in row:
                    if cell_val:
                        max_length = max(max_length, len(str(cell_val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 60)


def xlsx_generate(html: str, template: Optional[Dict] = None, metadata: Optional[Dict] = None) -> bytes:
    gen = XLSXGenerator()
    return gen.generate(html, template, metadata)
