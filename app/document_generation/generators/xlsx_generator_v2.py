import io
import re
from typing import Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.document_generation.generators.base import BaseGenerator
from app.document_generation.markdown_ast import parse, extract_text


class XLSXGeneratorV2(BaseGenerator):
    @property
    def format(self) -> str:
        return "xlsx-v2"

    @property
    def mime_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    @property
    def file_extension(self) -> str:
        return ".xlsx"

    def generate(self, html: str, template: Optional[Dict[str, Any]] = None, metadata: Optional[Dict[str, Any]] = None) -> bytes:
        return self._generate_v2(html, template, metadata)

    def generate_preview(self, html: str, template: Optional[Dict[str, Any]] = None) -> dict:
        ast = parse(html)
        tables = [b for b in ast if b["type"] == "table"]
        if tables:
            sheets = []
            for tbl in tables:
                rows_data = self._collect_table_rows(tbl)
                headers = rows_data[0] if rows_data else []
                body = rows_data[1:] if len(rows_data) > 1 else []
                sheets.append({
                    "name": extract_text(tbl.get("children", []))[:31] or "Sheet",
                    "headers": headers,
                    "rows": body[:100],
                    "total_rows": len(body),
                })
            return {"sheets": sheets}
        text_parts = []
        for b in ast:
            if b["type"] in ("paragraph", "heading"):
                text_parts.append(extract_text(b.get("children", [])))
        text = "\n".join(text_parts)[:500]
        return {
            "sheets": [{
                "name": "Content",
                "headers": ["Content"],
                "rows": [[line] for line in text.split("\n") if line.strip()],
                "total_rows": len(text.split("\n")),
            }]
        }

    def _generate_v2(self, html: str, template: Optional[Dict] = None, metadata: Optional[Dict] = None) -> bytes:
        wb = Workbook()
        wb.remove(wb.active)
        ast = parse(html)
        tables = [b for b in ast if b["type"] == "table"]

        if tables:
            for i, tbl in enumerate(tables):
                ws = wb.create_sheet(title=f"Table {i + 1}")
                self._write_table_to_sheet(ws, tbl)
        else:
            ws = wb.create_sheet(title="Content")
            self._write_content_to_sheet(ws, ast)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _collect_table_rows(self, tbl: dict) -> list[list[str]]:
        rows_data = []
        for child in tbl.get("children", []):
            for row in child.get("children", []):
                cells = [extract_text(c.get("children", [])) for c in row.get("children", [])]
                rows_data.append(cells)
        return rows_data

    def _write_table_to_sheet(self, ws, tbl: dict):
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        rows_data = self._collect_table_rows(tbl)
        if not rows_data:
            return

        for ri, row_data in enumerate(rows_data):
            for ci, cell_text in enumerate(row_data):
                cell = ws.cell(row=ri + 1, column=ci + 1, value=cell_text)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if ri == 0:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                elif ri % 2 == 0:
                    cell.fill = alt_fill

        num_cols = max(len(r) for r in rows_data) if rows_data else 1
        for ci in range(1, num_cols + 1):
            max_len = 0
            for row in ws.iter_rows(min_col=ci, max_col=ci, values_only=True):
                for val in row:
                    if val:
                        max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 60)

    def _write_content_to_sheet(self, ws, ast: list):
        heading_font = Font(bold=True, color="FFFFFF", size=13)
        heading_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        text_font = Font(size=10, color="333333")
        code_font = Font(name="Courier New", size=9, color="2D2D2D")
        code_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )
        wrap_align = Alignment(vertical="top", wrap_text=True)

        row_idx = 1
        for block in ast:
            if row_idx > 200:
                break
            if block["type"] == "heading":
                text = extract_text(block.get("children", []))
                cell = ws.cell(row=row_idx, column=1, value=text)
                cell.font = heading_font
                cell.fill = heading_fill
                cell.alignment = Alignment(vertical="center")
                cell.border = thin_border
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
                row_idx += 1
            elif block["type"] == "paragraph":
                text = extract_text(block.get("children", []))
                if text:
                    cell = ws.cell(row=row_idx, column=1, value=text)
                    cell.font = text_font
                    cell.alignment = wrap_align
                    cell.border = thin_border
                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
                    row_idx += 1
            elif block["type"] == "block_code":
                code = block.get("raw", "")
                for line in code.split("\n"):
                    if line.strip() or row_idx < 3:
                        cell = ws.cell(row=row_idx, column=1, value=line if line.strip() else "")
                        cell.font = code_font
                        cell.fill = code_fill
                        cell.alignment = wrap_align
                        cell.border = thin_border
                        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
                        row_idx += 1
            elif block["type"] == "list":
                for item in block.get("children", []):
                    text = extract_text(item.get("children", []))
                    if text:
                        bullet_cell = ws.cell(row=row_idx, column=1, value="•")
                        bullet_cell.font = text_font
                        bullet_cell.alignment = Alignment(horizontal="center", vertical="top")
                        bullet_cell.border = thin_border
                        val_cell = ws.cell(row=row_idx, column=2, value=text)
                        val_cell.font = text_font
                        val_cell.alignment = wrap_align
                        val_cell.border = thin_border
                        row_idx += 1

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 70


def xlsx_generate_v2(html: str, template: Optional[Dict] = None, metadata: Optional[Dict] = None) -> bytes:
    gen = XLSXGeneratorV2()
    return gen.generate(html, template, metadata)
